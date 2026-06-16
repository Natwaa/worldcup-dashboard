import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# โฟลเดอร์เก็บไฟล์ทายผลของผู้เล่น
UPLOAD_FOLDER = "uploads"

# ไฟล์ผลการแข่งขันจริง
RESULT_FILE = "data/results.csv"


def get_result(actual_home, actual_away, pred_home, pred_away):
    """
    กติกา:
    - ทายผลแพ้/ชนะ/เสมอถูก = 3 คะแนน
    - ทายสกอร์เจ้าบ้านถูก = 2 คะแนน
    - ทายสกอร์ทีมเยือนถูก = 2 คะแนน

    คะแนนเต็ม 7 คะแนน
    """

    if pred_home is None or pred_away is None:
        return 0

    score = 0

    # 1) ทายผลการแข่งขันถูก
    actual_result = (
        1 if actual_home > actual_away
        else -1 if actual_home < actual_away
        else 0
    )

    pred_result = (
        1 if pred_home > pred_away
        else -1 if pred_home < pred_away
        else 0
    )

    if actual_result == pred_result:
        score += 3

    # 2) ทายสกอร์เจ้าบ้านถูก
    if actual_home == pred_home:
        score += 2

    # 3) ทายสกอร์ทีมเยือนถูก
    if actual_away == pred_away:
        score += 2

    return score


def load_players():
    """
    อ่านไฟล์ Excel ของผู้เล่นทุกคน
    คำนวณคะแนนรวม
    และส่งกลับ Ranking
    """

    # อ่านผลการแข่งขันจริง
    results = pd.read_csv(
        RESULT_FILE,
        dtype={"match_no": int}
    )

    ranking = []

    # วนอ่านไฟล์ Excel ทุกไฟล์ใน uploads
    for file in Path(UPLOAD_FOLDER).glob("*.xlsx"):

        wb = load_workbook(file, data_only=True)

        # เลือกชีตตารางทายผล
        if "ตารางทายผล" in wb.sheetnames:
            ws = wb["ตารางทายผล"]
        else:
            ws = wb.active

        # ข้อมูลผู้เล่น
        player_name = ws["B1"].value
        champion_pick = ws["B2"].value

        total_points = 0

        # อ่านคู่แข่งขันทั้งหมด
        for row in range(5, 100):

            try:

                # Match Number
                match_no = ws[f"B{row}"].value

                if match_no is None:
                    continue

                match_no = int(match_no)

                # สกอร์ที่ผู้เล่นทาย
                pred_home = ws[f"E{row}"].value
                pred_away = ws[f"F{row}"].value

                # ค้นหาผลจริงของคู่นี้
                result_row = results[
                    results["match_no"] == match_no
                ]

                if len(result_row) == 0:
                    continue

                result_row = result_row.iloc[0]

                # คิดคะแนนเฉพาะคู่ที่แข่งจบแล้ว
                if result_row["status"] != "FINISHED":
                    continue

                score = get_result(
                    int(result_row["home_score"]),
                    int(result_row["away_score"]),
                    pred_home,
                    pred_away
                )

                total_points += score

            except Exception:
                # ข้ามแถวที่มีข้อมูลผิดรูปแบบ
                continue

        ranking.append({
            "name": player_name,
            "champion": champion_pick,
            "points": total_points
        })

    # เรียงคะแนนจากมากไปน้อย
    ranking.sort(
        key=lambda x: x["points"],
        reverse=True
    )

    for i, player in enumerate(ranking, start=1):
        player["rank"] = i

    return ranking